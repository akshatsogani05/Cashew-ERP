from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import jwt
import bcrypt
import random
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors as rl_colors

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

mongo_url = os.environ['MONGO_URL']
print("MONGO_URL =", mongo_url)
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Cashew ERP")
api = APIRouter(prefix="/api")

JWT_ALGO = "HS256"

ROLES = ["admin", "sales_manager", "production_manager", "inventory_manager", "accounts_manager", "dispatch_manager"]

WHOLE_GRADES = ["W180", "W210", "W240", "W320", "W450"]
PIECE_GRADES = ["Large Pieces", "Small Pieces", "Splits", "Butts", "Baby Bits", "Powder"]

# ============== Helpers ==============
def gen_id() -> str:
    return str(uuid.uuid4())

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {"sub": user_id, "email": email, "role": role,
               "exp": datetime.now(timezone.utc) + timedelta(hours=12), "type": "access"}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGO)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGO])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

def require_roles(*roles):
    async def checker(user: dict = Depends(get_current_user)):
        if user["role"] != "admin" and user["role"] not in roles:
            raise HTTPException(403, "Insufficient permissions")
        return user
    return checker

# ============== Models ==============
class LoginIn(BaseModel):
    email: EmailStr
    password: str

class GenericDoc(BaseModel):
    pass

# ============== Auth ==============
@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "Invalid email or password")
    token = create_access_token(user["id"], user["email"], user["role"])
    response.set_cookie("access_token", token, httponly=True, secure=False, samesite="lax", max_age=43200, path="/")
    return {"id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"], "token": token}

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user

@api.get("/auth/users")
async def list_users(user: dict = Depends(require_roles())):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return users

# ============== Generic CRUD factory ==============
def make_crud(collection: str, prefix: str, allowed_roles: List[str] = None):
    @api.get(f"/{prefix}")
    async def list_items(user: dict = Depends(get_current_user)):
        items = await db[collection].find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
        return items

    @api.post(f"/{prefix}")
    async def create_item(body: Dict[str, Any], user: dict = Depends(get_current_user)):
        doc = {**body, "id": gen_id(), "created_at": now_iso(), "created_by": user["email"]}
        await db[collection].insert_one(doc)
        doc.pop("_id", None)
        return doc

    @api.put(f"/{prefix}/{{item_id}}")
    async def update_item(item_id: str, body: Dict[str, Any], user: dict = Depends(get_current_user)):
        body.pop("id", None); body.pop("_id", None)
        body["updated_at"] = now_iso()
        await db[collection].update_one({"id": item_id}, {"$set": body})
        item = await db[collection].find_one({"id": item_id}, {"_id": 0})
        if not item:
            raise HTTPException(404, "Not found")
        return item

    @api.delete(f"/{prefix}/{{item_id}}")
    async def delete_item(item_id: str, user: dict = Depends(get_current_user)):
        await db[collection].delete_one({"id": item_id})
        return {"ok": True}

# Master data CRUD
make_crud("customers", "customers")
make_crud("suppliers", "suppliers")
make_crud("products", "products")
make_crud("categories", "categories")
make_crud("employees", "employees")
make_crud("warehouses", "warehouses")
make_crud("procurements", "procurements")
make_crud("inventory", "inventory")
make_crud("stock_movements", "stock-movements")
make_crud("production_batches", "production")
make_crud("sales_orders", "sales")
make_crud("dispatches", "dispatches")
make_crud("transactions", "accounts")
make_crud("notifications", "notifications")

# ============== Dashboard ==============
@api.get("/dashboard/summary")
async def dashboard_summary(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    month_start = today.replace(day=1).isoformat()
    year_start = today.replace(month=1, day=1).isoformat()
    today_iso = today.isoformat()

    # Today
    today_sales = await db.sales_orders.aggregate([
        {"$match": {"order_date": today_iso}},
        {"$group": {"_id": None, "total": {"$sum": "$total_value"}}}
    ]).to_list(1)
    today_production = await db.production_batches.aggregate([
        {"$match": {"start_date": today_iso}},
        {"$group": {"_id": None, "total": {"$sum": "$output_kg"}}}
    ]).to_list(1)
    today_dispatch_count = await db.dispatches.count_documents({"dispatch_date": today_iso})
    inv_items = await db.inventory.find({}, {"_id": 0}).to_list(2000)
    inventory_value = sum((i.get("current_stock", 0) * i.get("unit_value", 0)) for i in inv_items)

    # Month
    month_sales = await db.sales_orders.aggregate([
        {"$match": {"order_date": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_value"}, "count": {"$sum": 1}}}
    ]).to_list(1)
    month_prod = await db.production_batches.aggregate([
        {"$match": {"start_date": {"$gte": month_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$output_kg"}}}
    ]).to_list(1)
    month_revenue = month_sales[0]["total"] if month_sales else 0
    # estimate profit 25%
    month_profit = month_revenue * 0.25

    # YTD
    ytd_sales = await db.sales_orders.aggregate([
        {"$match": {"order_date": {"$gte": year_start}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_value"}}}
    ]).to_list(1)
    ytd_revenue = ytd_sales[0]["total"] if ytd_sales else 0
    ytd_profit = ytd_revenue * 0.25
    # growth: compare current month with previous month
    prev_month_dt = today.replace(day=1) - timedelta(days=1)
    prev_month_start = prev_month_dt.replace(day=1).isoformat()
    prev_month_end = today.replace(day=1).isoformat()
    prev_sales = await db.sales_orders.aggregate([
        {"$match": {"order_date": {"$gte": prev_month_start, "$lt": prev_month_end}}},
        {"$group": {"_id": None, "total": {"$sum": "$total_value"}}}
    ]).to_list(1)
    prev_rev = prev_sales[0]["total"] if prev_sales else 0
    growth = ((month_revenue - prev_rev) / prev_rev * 100) if prev_rev > 0 else 0

    return {
        "today": {
            "production_kg": today_production[0]["total"] if today_production else 0,
            "dispatch_count": today_dispatch_count,
            "sales_value": today_sales[0]["total"] if today_sales else 0,
            "inventory_value": round(inventory_value, 2),
        },
        "month": {
            "revenue": round(month_revenue, 2),
            "production_kg": month_prod[0]["total"] if month_prod else 0,
            "profit": round(month_profit, 2),
            "orders": month_sales[0]["count"] if month_sales else 0,
        },
        "ytd": {
            "revenue": round(ytd_revenue, 2),
            "profit": round(ytd_profit, 2),
            "growth_pct": round(growth, 1),
        },
    }

@api.get("/dashboard/trends")
async def dashboard_trends(user: dict = Depends(get_current_user)):
    # monthly buckets last 12 months
    today = datetime.now(timezone.utc).date()
    months = []
    for i in range(11, -1, -1):
        m = (today.replace(day=1) - timedelta(days=30 * i)).replace(day=1)
        months.append(m)
    months = sorted(set(months))[-12:]

    def in_month(d_iso: str, m):
        try:
            d = datetime.fromisoformat(d_iso).date()
            return d.year == m.year and d.month == m.month
        except Exception:
            return False

    sales = await db.sales_orders.find({}, {"_id": 0}).to_list(5000)
    prods = await db.production_batches.find({}, {"_id": 0}).to_list(5000)
    procs = await db.procurements.find({}, {"_id": 0}).to_list(5000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(5000)

    series = []
    for m in months:
        label = m.strftime("%b %y")
        s = sum(x.get("total_value", 0) for x in sales if in_month(x.get("order_date", ""), m))
        p = sum(x.get("output_kg", 0) for x in prods if in_month(x.get("start_date", ""), m))
        c = sum(x.get("total_cost", 0) for x in procs if in_month(x.get("purchase_date", ""), m))
        new_cust = sum(1 for x in customers if in_month(x.get("created_at", ""), m))
        profit = s * 0.25
        series.append({"month": label, "sales": round(s, 2), "production": round(p, 2),
                       "procurement": round(c, 2), "profit": round(profit, 2),
                       "new_customers": new_cust})
    return {"series": series}

# ============== Analytics ==============
@api.get("/analytics/rankings")
async def rankings(user: dict = Depends(get_current_user)):
    top_customers = await db.sales_orders.aggregate([
        {"$group": {"_id": "$customer_name", "revenue": {"$sum": "$total_value"}, "orders": {"$sum": 1}}},
        {"$sort": {"revenue": -1}}, {"$limit": 10}
    ]).to_list(10)
    top_products = await db.sales_orders.aggregate([
        {"$group": {"_id": "$product_name", "qty": {"$sum": "$quantity_kg"}, "revenue": {"$sum": "$total_value"}}},
        {"$sort": {"revenue": -1}}, {"$limit": 10}
    ]).to_list(10)
    top_suppliers = await db.procurements.aggregate([
        {"$group": {"_id": "$supplier_name", "qty": {"$sum": "$quantity_kg"}, "value": {"$sum": "$total_cost"}}},
        {"$sort": {"value": -1}}, {"$limit": 10}
    ]).to_list(10)
    return {"top_customers": top_customers, "top_products": top_products, "top_suppliers": top_suppliers}

@api.get("/analytics/kpis")
async def kpis(user: dict = Depends(get_current_user)):
    prods = await db.production_batches.find({}, {"_id": 0}).to_list(5000)
    sales = await db.sales_orders.find({}, {"_id": 0}).to_list(5000)
    customers = await db.customers.find({}, {"_id": 0}).to_list(5000)
    inventory = await db.inventory.find({}, {"_id": 0}).to_list(5000)

    avg_recovery = (sum(p.get("recovery_pct", 0) for p in prods) / len(prods)) if prods else 0
    avg_yield = (sum(p.get("yield_pct", 0) for p in prods) / len(prods)) if prods else 0
    avg_loss = (sum(p.get("loss_pct", 0) for p in prods) / len(prods)) if prods else 0
    total_input = sum(p.get("input_kg", 0) for p in prods)
    capacity = 100000  # kg monthly capacity
    capacity_util = (total_input / (capacity * 12)) * 100 if prods else 0

    total_sales_value = sum(s.get("total_value", 0) for s in sales)
    total_sales_kg = sum(s.get("quantity_kg", 0) for s in sales)
    asp = (total_sales_value / total_sales_kg) if total_sales_kg else 0

    # Repeat customers
    cust_orders = {}
    for s in sales:
        c = s.get("customer_name", "")
        cust_orders[c] = cust_orders.get(c, 0) + 1
    repeat = sum(1 for v in cust_orders.values() if v > 1)
    retention = (repeat / len(cust_orders) * 100) if cust_orders else 0

    inv_total = sum(i.get("current_stock", 0) * i.get("unit_value", 0) for i in inventory)
    turnover = (total_sales_value / inv_total) if inv_total else 0

    return {
        "avg_recovery_pct": round(avg_recovery, 2),
        "avg_yield_pct": round(avg_yield, 2),
        "avg_loss_pct": round(avg_loss, 2),
        "capacity_utilization_pct": round(min(capacity_util, 100), 2),
        "average_selling_price": round(asp, 2),
        "customer_retention_pct": round(retention, 2),
        "inventory_turnover": round(turnover, 2),
        "total_customers": len(customers),
        "repeat_customers": repeat,
        "total_orders": len(sales),
    }

# ============== Factory Health Score ==============
@api.get("/health-score")
async def health_score(user: dict = Depends(get_current_user)):
    prods = await db.production_batches.find({}, {"_id": 0}).to_list(5000)
    sales = await db.sales_orders.find({}, {"_id": 0}).to_list(5000)
    inv = await db.inventory.find({}, {"_id": 0}).to_list(5000)
    dispatches = await db.dispatches.find({}, {"_id": 0}).to_list(5000)

    checks = []

    # Recovery
    avg_recovery = (sum(p.get("recovery_pct", 0) for p in prods) / len(prods)) if prods else 0
    if avg_recovery >= 22: status = "green"; msg = f"Avg recovery {avg_recovery:.1f}% is healthy"
    elif avg_recovery >= 18: status = "yellow"; msg = f"Avg recovery {avg_recovery:.1f}% is below ideal"
    else: status = "red"; msg = f"Avg recovery {avg_recovery:.1f}% is critically low"
    checks.append({"key": "recovery", "label": "Recovery %", "status": status, "message": msg, "value": round(avg_recovery, 2)})

    # Production loss
    avg_loss = (sum(p.get("loss_pct", 0) for p in prods) / len(prods)) if prods else 0
    if avg_loss <= 3: status = "green"; msg = f"Production loss {avg_loss:.1f}% within target"
    elif avg_loss <= 6: status = "yellow"; msg = f"Production loss {avg_loss:.1f}% elevated"
    else: status = "red"; msg = f"Production loss {avg_loss:.1f}% critical"
    checks.append({"key": "loss", "label": "Production Loss", "status": status, "message": msg, "value": round(avg_loss, 2)})

    # Sales trend (last vs prev month)
    today = datetime.now(timezone.utc).date()
    ms = today.replace(day=1)
    pm_end = ms
    pm_start = (ms - timedelta(days=1)).replace(day=1)
    def in_range(iso, a, b):
        try:
            d = datetime.fromisoformat(iso).date()
            return a <= d < b
        except Exception:
            return False
    cur_sales = sum(s.get("total_value", 0) for s in sales if in_range(s.get("order_date", ""), ms, today + timedelta(days=1)))
    prev_sales = sum(s.get("total_value", 0) for s in sales if in_range(s.get("order_date", ""), pm_start, pm_end))
    if prev_sales == 0:
        sales_status = "yellow"; sales_msg = "Insufficient sales history"
    else:
        change = (cur_sales - prev_sales) / prev_sales * 100
        if change >= -5: sales_status = "green"; sales_msg = f"Sales trending {change:+.1f}% vs last month"
        elif change >= -15: sales_status = "yellow"; sales_msg = f"Sales declining {change:.1f}% vs last month"
        else: sales_status = "red"; sales_msg = f"Sales sharply declining {change:.1f}%"
    checks.append({"key": "sales_trend", "label": "Sales Trend", "status": sales_status, "message": sales_msg, "value": None})

    # Slow moving inventory: items with stock>0 and movement<3 in last 30 days
    movements = await db.stock_movements.find({}, {"_id": 0}).to_list(5000)
    cutoff = (today - timedelta(days=30)).isoformat()
    recent_moved = {m.get("item_id") for m in movements if m.get("date", "") >= cutoff}
    slow_count = sum(1 for i in inv if i.get("current_stock", 0) > 0 and i.get("id") not in recent_moved)
    if slow_count <= 2: status = "green"; msg = f"{slow_count} slow-moving items"
    elif slow_count <= 5: status = "yellow"; msg = f"{slow_count} slow-moving items, review pricing"
    else: status = "red"; msg = f"{slow_count} slow-moving items — action required"
    checks.append({"key": "slow_inventory", "label": "Slow Moving Inventory", "status": status, "message": msg, "value": slow_count})

    # Delayed dispatches
    delayed = sum(1 for d in dispatches if d.get("status") == "Delayed")
    if delayed == 0: status = "green"; msg = "All dispatches on schedule"
    elif delayed <= 2: status = "yellow"; msg = f"{delayed} delayed dispatches"
    else: status = "red"; msg = f"{delayed} delayed dispatches — operational issue"
    checks.append({"key": "delayed_dispatch", "label": "Delayed Dispatches", "status": status, "message": msg, "value": delayed})

    # Capacity utilization
    total_input = sum(p.get("input_kg", 0) for p in prods)
    cap_util = (total_input / (100000 * 12)) * 100 if prods else 0
    if cap_util >= 70: status = "green"; msg = f"Capacity utilization {cap_util:.1f}% strong"
    elif cap_util >= 40: status = "yellow"; msg = f"Capacity utilization {cap_util:.1f}% sub-optimal"
    else: status = "red"; msg = f"Capacity utilization {cap_util:.1f}% critically low"
    checks.append({"key": "capacity", "label": "Capacity Utilization", "status": status, "message": msg, "value": round(cap_util, 2)})

    # Customer concentration
    cust_rev = {}
    for s in sales:
        c = s.get("customer_name", "")
        cust_rev[c] = cust_rev.get(c, 0) + s.get("total_value", 0)
    total_rev = sum(cust_rev.values())
    if total_rev > 0:
        top_share = max(cust_rev.values()) / total_rev * 100
        if top_share <= 25: status = "green"; msg = f"Top customer share {top_share:.1f}% balanced"
        elif top_share <= 40: status = "yellow"; msg = f"Top customer share {top_share:.1f}% concentrated"
        else: status = "red"; msg = f"Top customer share {top_share:.1f}% — concentration risk"
    else:
        status = "yellow"; msg = "No sales data"; top_share = 0
    checks.append({"key": "concentration", "label": "Customer Concentration", "status": status, "message": msg, "value": round(top_share, 2)})

    # overall
    score_map = {"green": 100, "yellow": 60, "red": 20}
    overall = int(sum(score_map[c["status"]] for c in checks) / len(checks))
    if overall >= 80: overall_status = "green"
    elif overall >= 50: overall_status = "yellow"
    else: overall_status = "red"

    recommendations = []
    for c in checks:
        if c["status"] == "red":
            recommendations.append(f"Critical: address {c['label']} immediately. {c['message']}")
        elif c["status"] == "yellow":
            recommendations.append(f"Monitor {c['label']}: {c['message']}")

    return {"overall_score": overall, "overall_status": overall_status, "checks": checks, "recommendations": recommendations[:5]}

# ============== Reports (Excel / PDF) ==============
def df_to_xlsx(rows: List[dict], columns: List[str], title: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    ws.append([title])
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.append(columns)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="475569")
        cell.alignment = Alignment(horizontal="center")
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    for col in ws.columns:
        max_len = 12
        col_letter = None
        for cell in col:
            try:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len, 35)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def df_to_pdf(rows: List[dict], columns: List[str], title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
    styles = getSampleStyleSheet()
    elems = [Paragraph(f"<b>{title}</b>", styles["Title"]), Spacer(1, 12)]
    data = [columns] + [[str(r.get(c, "")) for c in columns] for r in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, rl_colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elems.append(table)
    doc.build(elems)
    buf.seek(0); return buf.getvalue()

REPORT_CONFIG = {
    "production": ("production_batches", ["batch_no", "start_date", "end_date", "input_kg", "output_kg", "recovery_pct", "yield_pct", "loss_pct"], "Production Report"),
    "sales": ("sales_orders", ["order_no", "order_date", "customer_name", "product_name", "quantity_kg", "rate", "total_value", "status"], "Sales Report"),
    "inventory": ("inventory", ["sku", "name", "category", "current_stock", "reserved_stock", "available_stock", "unit_value"], "Inventory Report"),
    "customers": ("customers", ["name", "country", "contact_person", "phone", "email", "credit_limit"], "Customer Report"),
    "suppliers": ("suppliers", ["name", "country", "contact_person", "phone", "email"], "Supplier Report"),
    "procurement": ("procurements", ["purchase_date", "supplier_name", "quantity_kg", "rate", "total_cost", "moisture_pct", "origin"], "Procurement Report"),
    "profit": ("sales_orders", ["order_no", "order_date", "customer_name", "total_value"], "Profit Report"),
}

@api.get("/reports/{report_type}")
async def get_report(report_type: str, fmt: str = Query("excel"), date_from: Optional[str] = None, date_to: Optional[str] = None, user: dict = Depends(get_current_user)):
    if report_type not in REPORT_CONFIG:
        raise HTTPException(404, "Unknown report")
    coll, cols, title = REPORT_CONFIG[report_type]
    query = {}
    date_field = {"production_batches": "start_date", "sales_orders": "order_date", "procurements": "purchase_date"}.get(coll)
    if date_field and (date_from or date_to):
        range_q = {}
        if date_from: range_q["$gte"] = date_from
        if date_to: range_q["$lte"] = date_to
        query[date_field] = range_q
    rows = await db[coll].find(query, {"_id": 0}).to_list(5000)
    if fmt == "pdf":
        data = df_to_pdf(rows, cols, title)
        return StreamingResponse(io.BytesIO(data), media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{report_type}_report.pdf"'})
    data = df_to_xlsx(rows, cols, title)
    return StreamingResponse(io.BytesIO(data), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{report_type}_report.xlsx"'})

# ============== Seeding ==============
async def seed_users():
    seeds = [
        ("admin@cashewerp.com", "admin123", "Aditi Sharma", "admin"),
        ("sales@cashewerp.com", "sales123", "Ravi Kumar", "sales_manager"),
        ("production@cashewerp.com", "prod123", "Meera Iyer", "production_manager"),
        ("inventory@cashewerp.com", "inv123", "Karthik Rao", "inventory_manager"),
        ("accounts@cashewerp.com", "acc123", "Priya Menon", "accounts_manager"),
        ("dispatch@cashewerp.com", "disp123", "Sundar Pillai", "dispatch_manager"),
    ]
    for email, pwd, name, role in seeds:
        existing = await db.users.find_one({"email": email})
        if not existing:
            await db.users.insert_one({"id": gen_id(), "email": email, "name": name, "role": role,
                                       "password_hash": hash_password(pwd), "created_at": now_iso()})
        else:
            # ensure password matches seed
            if not verify_password(pwd, existing["password_hash"]):
                await db.users.update_one({"email": email}, {"$set": {"password_hash": hash_password(pwd)}})

async def seed_master_and_transactions():
    if await db.customers.count_documents({}) > 0:
        return
    rnd = random.Random(42)
    countries = ["India", "UAE", "USA", "Germany", "UK", "Japan", "Netherlands", "Saudi Arabia"]
    cust_names = ["Premium Nuts Co", "Gulf Imports LLC", "Eurofoods Trading", "Tokyo Nuts KK",
                  "Atlantic Distributors", "Royal Foods Pvt", "Mediterranean Spice Co",
                  "Continental Snacks", "Nordic Imports", "Pacific Foods Ltd"]
    customers = []
    for i, n in enumerate(cust_names):
        c = {"id": gen_id(), "name": n, "country": rnd.choice(countries),
             "contact_person": f"Contact {i+1}", "phone": f"+91-90000{1000+i:04d}",
             "email": f"contact{i+1}@{n.lower().replace(' ', '')}.com",
             "credit_limit": rnd.choice([500000, 1000000, 2000000, 5000000]),
             "created_at": (datetime.now(timezone.utc) - timedelta(days=rnd.randint(30, 360))).isoformat()}
        customers.append(c)
    await db.customers.insert_many(customers)

    suppliers = []
    sup_names = ["Konkan Farms", "Goa Cashew Coop", "Mangalore Growers", "Vietnam Imports",
                 "Ivory Coast Traders", "Tanzania Cashews", "Kerala Plantations"]
    for i, n in enumerate(sup_names):
        suppliers.append({"id": gen_id(), "name": n, "country": rnd.choice(["India", "Vietnam", "Ivory Coast", "Tanzania"]),
                          "contact_person": f"Mgr {i+1}", "phone": f"+91-91111{2000+i:04d}",
                          "email": f"sales@{n.lower().replace(' ', '')}.com",
                          "created_at": now_iso()})
    await db.suppliers.insert_many(suppliers)

    cats = [{"id": gen_id(), "name": "Whole Grade", "created_at": now_iso()},
            {"id": gen_id(), "name": "Pieces", "created_at": now_iso()},
            {"id": gen_id(), "name": "Raw Material", "created_at": now_iso()},
            {"id": gen_id(), "name": "Packaging", "created_at": now_iso()}]
    await db.categories.insert_many(cats)

    products = []
    rates = {"W180": 1100, "W210": 950, "W240": 850, "W320": 750, "W450": 680,
             "Large Pieces": 600, "Small Pieces": 520, "Splits": 580, "Butts": 540,
             "Baby Bits": 450, "Powder": 380}
    for g in WHOLE_GRADES:
        products.append({"id": gen_id(), "sku": f"CSH-{g}", "name": g, "category": "Whole Grade",
                         "unit": "kg", "rate_per_kg": rates[g], "created_at": now_iso()})
    for g in PIECE_GRADES:
        products.append({"id": gen_id(), "sku": f"CSH-{g[:5].upper().replace(' ', '')}", "name": g, "category": "Pieces",
                         "unit": "kg", "rate_per_kg": rates[g], "created_at": now_iso()})
    products.append({"id": gen_id(), "sku": "RCN-RAW", "name": "Raw Cashew Nuts", "category": "Raw Material", "unit": "kg", "rate_per_kg": 180, "created_at": now_iso()})
    products.append({"id": gen_id(), "sku": "PKG-BAG", "name": "Vacuum Bags", "category": "Packaging", "unit": "pcs", "rate_per_kg": 8, "created_at": now_iso()})
    products.append({"id": gen_id(), "sku": "PKG-BOX", "name": "Corrugated Boxes", "category": "Packaging", "unit": "pcs", "rate_per_kg": 22, "created_at": now_iso()})
    products.append({"id": gen_id(), "sku": "PKG-LBL", "name": "Product Labels", "category": "Packaging", "unit": "pcs", "rate_per_kg": 1, "created_at": now_iso()})
    await db.products.insert_many(products)

    employees = []
    roles = ["Operator", "Supervisor", "QC Inspector", "Loader", "Accountant"]
    for i in range(15):
        employees.append({"id": gen_id(), "name": f"Employee {i+1}", "role": rnd.choice(roles),
                          "phone": f"+91-93333{i:04d}", "salary": rnd.randint(15000, 45000),
                          "join_date": (datetime.now(timezone.utc) - timedelta(days=rnd.randint(60, 1500))).date().isoformat(),
                          "created_at": now_iso()})
    await db.employees.insert_many(employees)

    warehouses = [
        {"id": gen_id(), "name": "Main Warehouse", "location": "Mangalore", "capacity_kg": 200000, "created_at": now_iso()},
        {"id": gen_id(), "name": "Cold Storage", "location": "Mangalore", "capacity_kg": 80000, "created_at": now_iso()},
        {"id": gen_id(), "name": "Dispatch Yard", "location": "Mangalore Port", "capacity_kg": 50000, "created_at": now_iso()},
    ]
    await db.warehouses.insert_many(warehouses)

    # Inventory
    inventory = []
    for p in products:
        stock = rnd.randint(500, 8000) if p["category"] != "Raw Material" else rnd.randint(20000, 60000)
        reserved = rnd.randint(0, stock // 4)
        inventory.append({"id": gen_id(), "sku": p["sku"], "name": p["name"], "category": p["category"],
                          "current_stock": stock, "reserved_stock": reserved,
                          "available_stock": stock - reserved, "unit_value": p["rate_per_kg"],
                          "batch_no": f"B{rnd.randint(10000, 99999)}",
                          "warehouse": rnd.choice([w["name"] for w in warehouses]),
                          "low_stock_threshold": 200,
                          "created_at": now_iso()})
    await db.inventory.insert_many(inventory)

    # 12 months of procurement, production, sales, dispatches
    today = datetime.now(timezone.utc).date()
    procurements = []
    productions = []
    sales = []
    dispatches = []
    movements = []
    transactions = []

    for m_offset in range(11, -1, -1):
        month_start = (today.replace(day=1) - timedelta(days=30 * m_offset)).replace(day=1)
        # 3-5 procurements per month
        for _ in range(rnd.randint(3, 6)):
            d = month_start + timedelta(days=rnd.randint(0, 27))
            qty = rnd.randint(8000, 25000)
            rate = rnd.randint(160, 210)
            freight = rnd.randint(5000, 18000)
            total = qty * rate
            procurements.append({"id": gen_id(),
                                 "purchase_date": d.isoformat(),
                                 "supplier_name": rnd.choice(sup_names),
                                 "quantity_kg": qty, "rate": rate, "total_cost": total,
                                 "freight_cost": freight, "landing_cost": total + freight,
                                 "moisture_pct": round(rnd.uniform(8, 12), 1),
                                 "origin": rnd.choice(["India", "Vietnam", "Ivory Coast", "Tanzania"]),
                                 "quality_notes": rnd.choice(["Grade A", "Mixed Quality", "Premium", "Standard"]),
                                 "created_at": d.isoformat()})

        # 4-7 production batches per month
        for b in range(rnd.randint(4, 8)):
            sd = month_start + timedelta(days=rnd.randint(0, 27))
            ed = sd + timedelta(days=rnd.randint(1, 4))
            input_kg = rnd.randint(2000, 8000)
            recovery = round(rnd.uniform(20, 24), 2)
            loss = round(rnd.uniform(1.5, 5.5), 2)
            output = round(input_kg * recovery / 100, 2)
            yield_pct = round(output / input_kg * 100, 2)
            # output by grade
            grade_split = {g: round(output * rnd.uniform(0.05, 0.25), 2) for g in WHOLE_GRADES}
            pieces_total = output - sum(grade_split.values())
            if pieces_total < 0: pieces_total = output * 0.1
            grade_split["Small Pieces"] = round(pieces_total * 0.4, 2)
            grade_split["Splits"] = round(pieces_total * 0.3, 2)
            grade_split["Powder"] = round(pieces_total * 0.3, 2)
            productions.append({"id": gen_id(),
                                "batch_no": f"BTH-{sd.strftime('%Y%m%d')}-{b+1:02d}",
                                "start_date": sd.isoformat(), "end_date": ed.isoformat(),
                                "input_kg": input_kg, "output_kg": output,
                                "recovery_pct": recovery, "yield_pct": yield_pct, "loss_pct": loss,
                                "grade_output": grade_split, "status": "Completed",
                                "created_at": sd.isoformat()})

        # 5-12 sales orders per month
        for o in range(rnd.randint(5, 12)):
            od = month_start + timedelta(days=rnd.randint(0, 27))
            cust = rnd.choice(customers)
            prod = rnd.choice([p for p in products if p["category"] in ["Whole Grade", "Pieces"]])
            qty = rnd.randint(200, 3000)
            rate = prod["rate_per_kg"] * rnd.uniform(0.95, 1.15)
            total = round(qty * rate, 2)
            status = rnd.choice(["Enquiry", "Quotation", "Order", "Production Allocation", "Dispatched", "Invoiced"])
            if m_offset > 0: status = rnd.choice(["Dispatched", "Invoiced", "Order"])
            sales.append({"id": gen_id(),
                          "order_no": f"SO-{od.strftime('%Y%m%d')}-{o+1:03d}",
                          "order_date": od.isoformat(),
                          "customer_name": cust["name"], "customer_id": cust["id"],
                          "product_name": prod["name"], "product_sku": prod["sku"],
                          "quantity_kg": qty, "rate": round(rate, 2), "total_value": total,
                          "delivery_date": (od + timedelta(days=rnd.randint(7, 30))).isoformat(),
                          "status": status, "created_at": od.isoformat()})

            if status in ["Dispatched", "Invoiced"]:
                disp_status = rnd.choice(["Delivered", "In Transit", "Delivered", "Delivered", "Delayed"])
                if m_offset > 1: disp_status = "Delivered"
                dispatches.append({"id": gen_id(),
                                   "dispatch_no": f"DP-{od.strftime('%Y%m%d')}-{o+1:03d}",
                                   "dispatch_date": (od + timedelta(days=rnd.randint(3, 14))).isoformat(),
                                   "order_no": f"SO-{od.strftime('%Y%m%d')}-{o+1:03d}",
                                   "vehicle_no": f"KA-19-{rnd.randint(1000, 9999)}",
                                   "lr_number": f"LR{rnd.randint(100000, 999999)}",
                                   "transporter": rnd.choice(["Speedy Logistics", "Maersk Line", "Indo Cargo", "Global Freight"]),
                                   "destination": cust["country"],
                                   "product_name": prod["name"],
                                   "quantity_kg": qty,
                                   "status": disp_status,
                                   "created_at": now_iso()})

            # transactions
            if status in ["Dispatched", "Invoiced"]:
                transactions.append({"id": gen_id(), "type": "Revenue",
                                     "description": f"Sale to {cust['name']} - {prod['name']}",
                                     "amount": total, "date": od.isoformat(),
                                     "category": "Sales", "created_at": now_iso()})
                # cost ~70%
                transactions.append({"id": gen_id(), "type": "Cost",
                                     "description": f"Raw material cost - {prod['name']}",
                                     "amount": -round(total * 0.55, 2), "date": od.isoformat(),
                                     "category": "Raw Material", "created_at": now_iso()})
                transactions.append({"id": gen_id(), "type": "Cost",
                                     "description": f"Labor cost - {prod['name']}",
                                     "amount": -round(total * 0.08, 2), "date": od.isoformat(),
                                     "category": "Labor", "created_at": now_iso()})
                transactions.append({"id": gen_id(), "type": "Cost",
                                     "description": f"Packaging cost",
                                     "amount": -round(total * 0.04, 2), "date": od.isoformat(),
                                     "category": "Packaging", "created_at": now_iso()})
                transactions.append({"id": gen_id(), "type": "Cost",
                                     "description": f"Freight",
                                     "amount": -round(total * 0.05, 2), "date": od.isoformat(),
                                     "category": "Freight", "created_at": now_iso()})
                transactions.append({"id": gen_id(), "type": "Cost",
                                     "description": f"Utilities",
                                     "amount": -round(total * 0.03, 2), "date": od.isoformat(),
                                     "category": "Utilities", "created_at": now_iso()})

    if procurements: await db.procurements.insert_many(procurements)
    if productions: await db.production_batches.insert_many(productions)
    if sales: await db.sales_orders.insert_many(sales)
    if dispatches: await db.dispatches.insert_many(dispatches)
    if transactions: await db.transactions.insert_many(transactions)

    # Notifications
    notes = []
    low_stock = [i for i in inventory if i["available_stock"] < 1000]
    for i in low_stock[:5]:
        notes.append({"id": gen_id(), "type": "low_stock", "severity": "warning",
                      "title": f"Low stock: {i['name']}",
                      "message": f"Only {i['available_stock']} {('kg' if 'PKG' not in i['sku'] else 'pcs')} available",
                      "read": False, "date": now_iso(), "created_at": now_iso()})
    pending = [s for s in sales if s["status"] in ["Enquiry", "Quotation"]][:5]
    for s in pending:
        notes.append({"id": gen_id(), "type": "pending_order", "severity": "info",
                      "title": f"Pending: {s['order_no']}",
                      "message": f"{s['customer_name']} - awaiting confirmation",
                      "read": False, "date": now_iso(), "created_at": now_iso()})
    delayed = [d for d in dispatches if d["status"] == "Delayed"][:3]
    for d in delayed:
        notes.append({"id": gen_id(), "type": "delayed_dispatch", "severity": "danger",
                      "title": f"Delayed: {d['dispatch_no']}",
                      "message": f"To {d['destination']} - immediate attention required",
                      "read": False, "date": now_iso(), "created_at": now_iso()})
    if notes: await db.notifications.insert_many(notes)

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await seed_users()
    await seed_master_and_transactions()
    logger.info("Cashew ERP seed complete")

# Include router
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "server:app",
        host="0.0.0.0",
        port=8000,
        reload=True
    )